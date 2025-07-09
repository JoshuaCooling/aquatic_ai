import cv2
import torch
import torchvision.transforms as transforms
from PIL import Image
from torchvision import models
import tkinter as tk
from tkinter import filedialog, Label, Button, StringVar, OptionMenu, Entry
import threading
import pandas as pd
import os
import openpyxl
import timm


# Define available models and their architectures
MODEL_PATHS = {
    "DenseNet121": 'invasive_species_model.pth',
    "EfficientNet": 'efficientnet_aquatic_plants.pth',
    "Swin Transformer": 'swin_tiny(74).pth',
    "Image Net Transformer": 'vision_transformer_model_copy.pth'
}

# Default selected model default Values
selected_model_name = "EfficientNet"  
confidence_threshold = 0.50
excel_file = "classification_results.xlsx"
results_dir = "classified_images"
os.makedirs(results_dir, exist_ok=True)

# Define image transformations
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
])

# Class labels
class_labels = ["bladderwort", "canadian_waterweed", "carolina_fanwort", "coontail", 
                "curly_leaf_pondweed", "eurasian_watermilfoil", "european_frogbit",
                "hydrilla", "parrotfeather", "richards_pondweed", "siberian_watermilfoil",
                "starry_stonewort"]

# Invasive species class labels
invasive_classes = ["carolina_fanwort","curly_leaf_pondweed", "eurasian_watermilfoil", 
                    "european_frogbit","hydrilla", "parrotfeather",
                    "starry_stonewort"]

# Function to load EfficientNet model
def load_efficientnet():
    model_path = "efficientnet_aquatic_plants.pth"
    model = torch.load(model_path, map_location=torch.device("cpu"), weights_only=False)
    model.eval()
    return model

# Function to load DenseNet model
def load_densenet():
    model = models.densenet121(weights=None)
    num_ftrs = model.classifier.in_features
    model.classifier = torch.nn.Linear(num_ftrs, 12)
    model.load_state_dict(torch.load("invasive_species_model.pth", map_location=torch.device('cpu')))
    model.eval()
    return model

def load_swin():
    model = timm.create_model('swin_tiny_patch4_window7_224', pretrained=False)
    # Path to your .pth checkpoint
    checkpoint_path = 'swin_tiny(74).pth'

    # Load the state dict
    checkpoint = torch.load(checkpoint_path, map_location='cpu')

    # If the checkpoint has a 'state_dict' key (common in training checkpoints)
    if 'state_dict' in checkpoint:
        checkpoint = checkpoint['state_dict']

    # Remove 'module.' prefix if the model was saved using DataParallel
    new_checkpoint = {}
    for k, v in checkpoint.items():
        if k.startswith('module.'):
            k = k[len('module.'):]
        new_checkpoint[k] = v

    # Load weights into the model
    model.load_state_dict(new_checkpoint, strict=False)

    model.eval()

    return model

def load_imageNet():
    model = timm.create_model('vit_base_patch16_224', pretrained=False)
    model.head = torch.nn.Linear(model.head.in_features, 12)
    model.load_state_dict(torch.load("vision_transformer_model_copy.pth"))

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model.to(device)
    model.eval()
    return model

# Load the selected model
if selected_model_name == "EfficientNet":
    model = load_efficientnet()
elif selected_model_name == "DenseNet121":
    model = load_densenet()

# Time conversion
def seconds_to_hms(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

# Function to classify an image frame
def classify_frame(frame, threshold=0.01):
    image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
    image = transform(image).unsqueeze(0)
    
    with torch.no_grad():
        outputs = model(image)
        probabilities = torch.nn.functional.softmax(outputs[0], dim=0)
    
    max_prob, predicted_class = torch.max(probabilities, 0)
    
    if max_prob.item() < threshold:
        return "Other", max_prob.item()
    return class_labels[predicted_class], max_prob.item()

# Function to save results to an excel sheet
def log_results(time, image_path, predicted_class, confidence):

    time_formatted = seconds_to_hms(time)

    new_data = pd.DataFrame([{
        "Time": time_formatted,
        "Predicted Class": predicted_class,
        "Confidence": round(confidence, 2),
        "Image Path": image_path
    }])

    try:
        if os.path.exists(excel_file):
            df = pd.read_excel(excel_file)
        else:
            df = pd.DataFrame(columns=["Time", "Predicted Class", "Confidence", "Image Path"])

        if df.empty or df.dropna(how="all").empty:
            df = new_data
        else:
            df = pd.concat([df, new_data], ignore_index=True)
        df.to_excel(excel_file, index=False)

        # Set column widths
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        column_widths = {
            "A": 10,
            "B": 25,
            "C": 10,
            "D": 35
        }
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
            wb.save(excel_file)

    except Exception as e:
        print(f"Error saving to Excel: {e}")

# Function to process images
def process_image(image_path):
    time = 0
    image = cv2.imread(image_path)
    predicted_class, confidence = classify_frame(image)
    result_label.config(text=f"Predicted: {predicted_class} ({confidence:.2f})")

    # Log the result
    log_results(time, image_path, predicted_class, confidence)

# Function to process video
def process_video(video_path, interval=1):  
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        result_label.config(text="Error: Could not open video.")
        return
    
    fps = int(cap.get(cv2.CAP_PROP_FPS))  
    frame_interval = fps * interval  
    frame_count = 0

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        frame_count += 1
        if frame_count % frame_interval == 0:  
            predicted_class, confidence = classify_frame(frame)
            print(f"Time {frame_count / fps:.2f}s: Predicted class: {predicted_class}, Confidence: {confidence:.2f}")

            # Save the frame as an image
            image_filename = os.path.join(results_dir, f"classified_{frame_count}.jpg")
            cv2.imwrite(image_filename, frame)
            time = frame_count / fps
            # Log results for each processed frame
            if predicted_class in invasive_classes and confidence >= confidence_threshold:
                log_results(time, image_filename, predicted_class, confidence)

        cv2.imshow("Video Frames", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    convert_paths_to_hyperlinks(excel_file)
    cap.release()
    cv2.destroyAllWindows()

# Function to upload a file
def upload_file(*args):
    file_path = filedialog.askopenfilename(filetypes=[("Image/Video", "*.jpg;*.jpeg;*.png;*.mp4")])

    os.makedirs(results_dir, exist_ok=True)

    if file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
        process_image(file_path)
    elif file_path.lower().endswith('.mp4'):
        threading.Thread(target=process_video, args=(file_path,)).start()
    else:
        result_label.config(text="Unsupported file format.")

# Function to switch between models
def update_model(*args):
    global model

    selected_model = model_var.get()

    if selected_model == "EfficientNet":
        model = load_efficientnet()
        print("Loaded EfficientNet")

    elif selected_model == "DenseNet121":
        model = load_densenet()
        print("Loaded DenseNet121")

    result_label.config(text=f"Switched to {selected_model}.")

# Function to update the Excel file name
def update_excel_filename(*args):
    global excel_file

    base_filename = excel_name_var.get().strip()

    if base_filename:
        excel_file = base_filename + ".xlsx"

    else:
        excel_file = "classification_results.xlsx"

    result_label.config(text=f"Excel file name set to: {excel_file}")

# Function to update the Excel file name
def update_folder_name(*args):
    global results_dir

    folder_name = folder_name_var.get().strip()

    if folder_name:
        results_dir = folder_name
    else:
        results_dir = "classified_images"

    result_label.config(text=f"Image folder name set to: {results_dir}")


# Function to update the confidence threshold from the input box
def update_confidence_threshold(*args):
    global confidence_threshold

    try:
        # Get the confidence value from the entry, default to 0.50 if invalid input
        confidence_threshold = float(confidence_entry.get())

        if confidence_threshold < 0 or confidence_threshold > 1:
            result_label.config(text="Confidence must be between 0 and 1.")
            confidence_threshold = 0.50

        else:
            result_label.config(text=f"Confidence threshold set to: {confidence_threshold:.2f}")

    except ValueError:
        result_label.config(text="Invalid confidence value. Please enter a number.")
        confidence_threshold = 0.50


# Add loop after excel sheet is made to convert image column to hyperlinks
def convert_paths_to_hyperlinks(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"D{row}"]
            image_path = cell.value
            if image_path and os.path.exists(image_path):
                cell.hyperlink = image_path
                cell.style = "Hyperlink"

        sheet.column_dimensions["D"].width = 50

        wb.save(excel_file)
        print("Hyperlinks added successfully.")

    except Exception as e:
        print(f"Error updating hyperlinks: {e}")


# GUI Setup
root = tk.Tk()
root.title("Underwater Invasive Species Classifier")
root.geometry("400x375")

# Model Selection
Label(root, text="Select Model:").pack()
model_var = StringVar(root)
model_var.set(selected_model_name)  
model_var.trace_add("write", update_model)  
model_menu = OptionMenu(root, model_var, *MODEL_PATHS.keys())
model_menu.pack(pady=5)

# excel file name
Label(root, text="Enter Excel File Name:").pack(pady=5)
excel_name_var = StringVar(root)
excel_name_var.set(excel_file.strip(".xlsx"))
excel_name_var.trace_add("write", update_excel_filename)
excel_name_entry = Entry(root, textvariable=excel_name_var)
excel_name_entry.pack(pady=5)

# Image folder file name
Label(root, text="Image Folder Name:").pack(pady=5)
folder_name_var = StringVar(root)
folder_name_var.set(results_dir)
folder_name_var.trace_add("write", update_folder_name)
folder_name_entry = Entry(root, textvariable=folder_name_var)
folder_name_entry.pack(pady=5)

# Confidence Selection
Label(root, text="Set Confidence Threshold (0-1):").pack(pady=5)
confidence_entry = tk.Entry(root)
confidence_entry.pack(pady=5)
confidence_entry.insert(0, str(confidence_threshold))
confidence_entry.bind("<KeyRelease>", update_confidence_threshold)

# Upload Button
Label(root, text="Upload an image or video").pack(pady=10)
Button(root, text="Upload File", command=upload_file).pack()
result_label = Label(root, text="")
result_label.pack(pady=20)

root.mainloop()
